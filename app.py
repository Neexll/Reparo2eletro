from flask import Flask, render_template, request, redirect, url_for, flash, get_flashed_messages, Response, jsonify
import sqlite3
from datetime import datetime
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import collections
from fpdf import FPDF
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Use Agg backend for non-interactive plotting

app = Flask(__name__)
app.secret_key = 'sua_chave_secreta_aqui'  # Necessário para sessões e mensagens flash

# Configuração do banco de dados
def init_db():
    with app.app_context():
        db = get_db()
        with app.open_resource('schema.sql', mode='r') as f:
            db.cursor().executescript(f.read())
        db.commit()

def get_db():
    db = sqlite3.connect(
        'database.db',
        detect_types=sqlite3.PARSE_DECLTYPES
    )
    db.row_factory = sqlite3.Row
    return db

def close_db(e=None):
    db = getattr(Flask, '_database', None)
    if db is not None:
        db.close()

@app.context_processor
def inject_now():
    return {'now': datetime.now()}

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/')
def index():
    db = get_db()
    search_numero = request.args.get('numero_pedido', '')

    query = '''
        SELECT p.id, p.numero_pedido, p.tecnico_nome, 
               strftime('%d/%m/%Y %H:%M', datetime(p.data_criacao, 'localtime')) as data_formatada,
               GROUP_CONCAT(i.nome, ', ') as itens_defeituosos
        FROM pedidos p
        LEFT JOIN pecas i ON p.id = i.pedido_id
    '''
    params = []

    if search_numero:
        query += " WHERE p.numero_pedido LIKE ?"
        params.append(f"%{search_numero}%")

    query += " GROUP BY p.id ORDER BY p.id DESC"

    pedidos = db.execute(query, params).fetchall()
    db.close()
    return render_template('index.html', pedidos=pedidos)

@app.route('/pedido/add', methods=['GET', 'POST'])
def add_pedido():
    if request.method == 'POST':
        numero_pedido = request.form.get('numero_pedido', '').strip()
        tecnico_nome = request.form.get('tecnico_nome', '').strip()

        if not numero_pedido or not tecnico_nome:
            flash('Número do pedido and nome do técnico são obrigatórios.', 'error')
            return redirect(url_for('add_pedido'))

        db = get_db()
        try:
            cursor = db.cursor()
            cursor.execute('INSERT INTO pedidos (numero_pedido, tecnico_nome) VALUES (?, ?)', (numero_pedido, tecnico_nome))
            pedido_id = cursor.lastrowid

            # Processar peças adicionadas
            nomes_pecas = request.form.getlist('peca_nome[]')
            defeitos_pecas = request.form.getlist('peca_defeito[]')

            for nome, defeito in zip(nomes_pecas, defeitos_pecas):
                if nome and defeito:
                    cursor.execute('INSERT INTO pecas (nome, defeito, pedido_id) VALUES (?, ?, ?)', (nome, defeito, pedido_id))

            db.commit()
            flash('Pedido e peças foram salvos com sucesso!', 'success')
            return redirect(url_for('view_pedido', id=pedido_id))
        except sqlite3.Error as e:
            flash(f'Erro ao criar pedido: {e}', 'error')
        finally:
            db.close()
        return redirect(url_for('add_pedido'))

    db = get_db()
    tecnicos = db.execute('SELECT nome FROM tecnicos ORDER BY nome').fetchall()
    tipos_pecas = db.execute('SELECT id, nome FROM tipos_pecas ORDER BY nome').fetchall()
    db.close()
    return render_template('add_pedido.html', tecnicos=tecnicos, tipos_pecas=tipos_pecas)

@app.route('/pedido/<int:id>/edit', methods=['GET', 'POST'])
def edit_pedido(id):
    db = get_db()
    pedido = db.execute('SELECT * FROM pedidos WHERE id = ?', (id,)).fetchone()
    if not pedido:
        flash('Pedido não encontrado.', 'error')
        return redirect(url_for('index'))

    if request.method == 'POST':
        numero_pedido = request.form.get('numero_pedido', '').strip()
        tecnico_nome = request.form.get('tecnico_nome', '').strip()

        if not numero_pedido or not tecnico_nome:
            flash('Número do pedido e nome do técnico são obrigatórios.', 'error')
            return redirect(url_for('edit_pedido', id=id))

        try:
            cursor = db.cursor()
            # Atualiza os detalhes do pedido
            cursor.execute('UPDATE pedidos SET numero_pedido = ?, tecnico_nome = ? WHERE id = ?', (numero_pedido, tecnico_nome, id))

            # Deleta as peças antigas para reinserir as novas (abordagem mais simples)
            cursor.execute('DELETE FROM pecas WHERE pedido_id = ?', (id,))

            # Processa as peças enviadas
            nomes_pecas = request.form.getlist('peca_nome[]')
            defeitos_pecas = request.form.getlist('peca_defeito[]')

            for nome, defeito in zip(nomes_pecas, defeitos_pecas):
                if nome and defeito:
                    cursor.execute('INSERT INTO pecas (nome, defeito, pedido_id) VALUES (?, ?, ?)', (nome, defeito, id))

            db.commit()
            flash('Pedido atualizado com sucesso!', 'success')
            return redirect(url_for('view_pedido', id=id))
        except sqlite3.Error as e:
            flash(f'Erro ao atualizar pedido: {e}', 'error')
        finally:
            db.close()
        return redirect(url_for('edit_pedido', id=id))

    # Método GET
    pecas = db.execute('SELECT * FROM pecas WHERE pedido_id = ?', (id,)).fetchall()
    tecnicos = db.execute('SELECT nome FROM tecnicos ORDER BY nome').fetchall()
    tipos_pecas = db.execute('SELECT id, nome FROM tipos_pecas ORDER BY nome').fetchall()
    db.close()
    return render_template('edit_pedido.html', pedido=pedido, pecas=pecas, tecnicos=tecnicos, tipos_pecas=tipos_pecas)

@app.route('/pedido/<int:id>')
def view_pedido(id):
    db = get_db()
    pedido = db.execute('SELECT * FROM pedidos WHERE id = ?', (id,)).fetchone()
    if not pedido:
        return 'Pedido não encontrado', 404
    
    pecas = db.execute('SELECT * FROM pecas WHERE pedido_id = ?', (id,)).fetchall()
    db.close()
    return render_template('view_pedido.html', pedido=pedido, pecas=pecas)

@app.route('/pedido/<int:pedido_id>/add_peca', methods=['POST'])
def add_peca_to_pedido(pedido_id):
    nome_peca = request.form.get('peca', '').strip()
    defeito = request.form.get('defeito', '').strip()

    if not nome_peca or not defeito:
        flash('Nome da peça e defeito são obrigatórios.', 'error')
        return redirect(url_for('view_pedido', id=pedido_id))

    try:
        db = get_db()
        db.execute('INSERT INTO pecas (nome, defeito, pedido_id) VALUES (?, ?, ?)', (nome_peca, defeito, pedido_id))
        db.commit()
        flash('Peça adicionada ao pedido com sucesso!', 'success')
    except sqlite3.Error as e:
        flash(f'Erro ao adicionar peça: {e}', 'error')
    finally:
        db.close()
    return redirect(url_for('view_pedido', id=pedido_id))

@app.route('/api/pecas/<int:peca_id>/defeitos')
def get_defeitos(peca_id):
    db = get_db()
    try:
        defeitos = db.execute(
            'SELECT descricao FROM defeitos_comuns WHERE tipo_peca_id = ? ORDER BY descricao',
            (peca_id,)
        ).fetchall()
        # Converte o resultado para uma lista de strings
        defeitos_lista = [defeito['descricao'] for defeito in defeitos]
        return jsonify(defeitos_lista)
    except Exception as e:
        print(f"Erro ao buscar defeitos: {e}")
        return jsonify([])
    finally:
        db.close()

@app.route('/add_peca_tipo', methods=['POST'])
def add_peca_tipo():
    nome_peca = request.form.get('nome_peca', '').strip()
    defeitos_str = request.form.get('defeitos', '').strip()

    if not nome_peca or not defeitos_str:
        flash('Nome da peça e pelo menos um defeito são obrigatórios.', 'error')
        return redirect(url_for('manage_pecas'))

    defeitos = [d.strip() for d in defeitos_str.split(',') if d.strip()]

    db = get_db()
    try:
        # Verifica se a peça já existe
        cursor = db.cursor()
        cursor.execute('SELECT id FROM tipos_pecas WHERE nome = ?', (nome_peca,))
        peca_existente = cursor.fetchone()

        if peca_existente:
            flash(f'O tipo de peça "{nome_peca}" já existe.', 'error')
        else:
            # Insere o novo tipo de peça
            cursor.execute('INSERT INTO tipos_pecas (nome) VALUES (?)', (nome_peca,))
            tipo_peca_id = cursor.lastrowid

            # Insere os defeitos associados
            for defeito in defeitos:
                cursor.execute('INSERT INTO defeitos_comuns (descricao, tipo_peca_id) VALUES (?, ?)', (defeito, tipo_peca_id))
            
            db.commit()
            flash(f'Novo tipo de peça "{nome_peca}" adicionado com sucesso!', 'success')

    except sqlite3.IntegrityError as e:
        db.rollback()
        flash(f'Erro ao adicionar a peça: {e}', 'error')
    finally:
        db.close()

    return redirect(url_for('manage_pecas'))

@app.route('/pedido/<int:id>/delete', methods=['POST'])
def delete_pedido(id):
    db = get_db()
    try:
        # Deleta primeiro as peças associadas ao pedido
        db.execute('DELETE FROM pecas WHERE pedido_id = ?', (id,))
        # Depois deleta o pedido principal
        db.execute('DELETE FROM pedidos WHERE id = ?', (id,))
        db.commit()
        flash('Pedido e todas as suas peças foram removidos com sucesso!', 'success')
    except sqlite3.Error as e:
        db.rollback()
        flash(f'Erro ao remover o pedido: {e}', 'error')
    finally:
        db.close()
    return redirect(url_for('index'))

@app.route('/manage_pecas')
def manage_pecas():
    db = get_db()
    tipos_pecas = db.execute('SELECT id, nome FROM tipos_pecas ORDER BY nome').fetchall()
    db.close()
    return render_template('manage_pecas.html', tipos_pecas=tipos_pecas)

@app.route('/edit_peca_tipo/<int:id>', methods=['GET', 'POST'])
def edit_peca_tipo(id):
    db = get_db()
    # Busca o tipo de peça para garantir que ele existe
    tipo_peca = db.execute('SELECT id, nome FROM tipos_pecas WHERE id = ?', (id,)).fetchone()
    if tipo_peca is None:
        flash('Tipo de peça não encontrado.', 'error')
        return redirect(url_for('manage_pecas'))

    if request.method == 'POST':
        nome_peca = request.form.get('nome_peca', '').strip()
        defeitos_str = request.form.get('defeitos', '').strip()

        if not nome_peca or not defeitos_str:
            flash('Nome da peça e pelo menos um defeito são obrigatórios.', 'error')
            # Se houver erro, recarrega a página com os dados atuais
            defeitos_rows = db.execute('SELECT descricao FROM defeitos_comuns WHERE tipo_peca_id = ?', (id,)).fetchall()
            defeitos_str_atual = ', '.join([d['descricao'] for d in defeitos_rows])
            return render_template('edit_peca_tipo.html', tipo_peca=tipo_peca, defeitos_str=defeitos_str_atual)

        defeitos_list = [d.strip() for d in defeitos_str.split(',') if d.strip()]

        try:
            cursor = db.cursor()
            # Atualiza o nome da peça
            cursor.execute('UPDATE tipos_pecas SET nome = ? WHERE id = ?', (nome_peca, id))
            # Deleta os defeitos antigos
            cursor.execute('DELETE FROM defeitos_comuns WHERE tipo_peca_id = ?', (id,))
            # Insere os novos defeitos
            for defeito in defeitos_list:
                cursor.execute('INSERT INTO defeitos_comuns (descricao, tipo_peca_id) VALUES (?, ?)', (defeito, id))
            db.commit()
            flash('Tipo de peça atualizado com sucesso!', 'success')
            return redirect(url_for('manage_pecas'))
        except sqlite3.Error as e:
            db.rollback()
            flash(f'Erro ao atualizar o tipo de peça: {e}', 'error')
        finally:
            db.close()
        return redirect(url_for('edit_peca_tipo', id=id))

    # Método GET: Carrega os dados para exibir no formulário
    defeitos_rows = db.execute('SELECT descricao FROM defeitos_comuns WHERE tipo_peca_id = ?', (id,)).fetchall()
    db.close()
    defeitos_str = ', '.join([d['descricao'] for d in defeitos_rows])
    
    return render_template('edit_peca_tipo.html', tipo_peca=tipo_peca, defeitos_str=defeitos_str)

@app.route('/delete_peca_tipo/<int:id>', methods=['POST'])
def delete_peca_tipo(id):
    db = get_db()
    try:
        # Em uma transação, primeiro deleta os defeitos e depois a peça
        cursor = db.cursor()
        cursor.execute('DELETE FROM defeitos_comuns WHERE tipo_peca_id = ?', (id,))
        cursor.execute('DELETE FROM tipos_pecas WHERE id = ?', (id,))
        db.commit()
        flash('Tipo de peça removido com sucesso!', 'success')
    except sqlite3.Error as e:
        db.rollback()
        flash(f'Erro ao remover o tipo de peça: {e}', 'error')
    finally:
        db.close()
    return redirect(url_for('manage_pecas'))

# --- Rotas para Gerenciamento de Técnicos ---

@app.route('/manage_tecnicos')
def manage_tecnicos():
    db = get_db()
    tecnicos = db.execute('SELECT id, nome FROM tecnicos ORDER BY nome').fetchall()
    db.close()
    return render_template('manage_tecnicos.html', tecnicos=tecnicos)

@app.route('/add_tecnico', methods=['POST'])
def add_tecnico():
    nome = request.form.get('nome', '').strip()
    if nome:
        try:
            db = get_db()
            db.execute('INSERT INTO tecnicos (nome) VALUES (?)', (nome,))
            db.commit()
            flash('Técnico adicionado com sucesso!', 'success')
        except sqlite3.IntegrityError:
            flash('Este técnico já existe.', 'error')
        finally:
            db.close()
    else:
        flash('O nome do técnico é obrigatório.', 'error')
    return redirect(url_for('manage_tecnicos'))

@app.route('/delete_tecnico/<int:id>', methods=['POST'])
def delete_tecnico(id):
    try:
        db = get_db()
        db.execute('DELETE FROM tecnicos WHERE id = ?', (id,))
        db.commit()
        flash('Técnico removido com sucesso!', 'success')
    except sqlite3.Error as e:
        flash(f'Erro ao remover técnico: {e}', 'error')
    finally:
        db.close()
    return redirect(url_for('manage_tecnicos'))

@app.route('/export')
def export():
    db = get_db()

    # 1. Coleta de dados
    pecas_list = db.execute('''
        SELECT p.numero_pedido, p.tecnico_nome, i.nome, i.defeito, STRFTIME('%d/%m/%Y', p.data_criacao) as data_formatada
        FROM pedidos p JOIN pecas i ON p.id = i.pedido_id
        ORDER BY p.id DESC
    ''').fetchall()
    pie_chart_data = db.execute('SELECT nome || \' - \' || defeito as label, COUNT(*) as total FROM pecas GROUP BY label ORDER BY total DESC').fetchall()
    bar_chart_stats = db.execute('''
        SELECT strftime('%Y-%m', datetime(p.data_criacao, 'localtime')) as mes, i.nome as peca_nome, COUNT(i.id) as total
        FROM pedidos p JOIN pecas i ON p.id = i.pedido_id
        GROUP BY mes, peca_nome ORDER BY mes, peca_nome
    ''').fetchall()
    db.close()

    # 2. Criação dos gráficos como imagens
    pie_chart_img = create_pie_chart_image(pie_chart_data)
    bar_chart_img = create_bar_chart_image(bar_chart_stats)

    # 3. Criação da planilha Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    # Título
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = 'Relatório de Peças e Estatísticas'
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    # Inserir gráficos
    if pie_chart_img:
        img = Image(pie_chart_img)
        img.width, img.height = 400, 300
        ws.add_image(img, 'A3')
    if bar_chart_img:
        img = Image(bar_chart_img)
        img.width, img.height = 400, 300
        ws.add_image(img, 'E3')

    # Tabela de dados (começando mais para baixo)
    table_start_row = 20
    ws.cell(row=table_start_row, column=1, value='Lista de Peças').font = Font(size=12, bold=True)

    headers = ['Pedido', 'Defeito', 'Peça', 'Técnico', 'Data']
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=table_start_row + 1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill

    for peca in pecas_list:
        ws.append([peca['numero_pedido'], peca['defeito'], peca['nome'], peca['tecnico_nome'], peca['data_formatada']])

    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # --- Finalização ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment;filename=relatorio_completo_com_graficos.xlsx"}
    )

@app.route('/export/pdf')
def export_pdf():
    db = get_db()

    # 1. Dados para a lista de peças
    pecas_list = db.execute('''
        SELECT p.numero_pedido, p.tecnico_nome, i.nome, i.defeito, STRFTIME('%d/%m/%Y', datetime(p.data_criacao, 'localtime')) as data_formatada
        FROM pedidos p JOIN pecas i ON p.id = i.pedido_id
        ORDER BY p.id DESC
    ''').fetchall()

    # 2. Dados para o gráfico de pizza
    pie_chart_data = db.execute('''
        SELECT nome || ' - ' || defeito as label, COUNT(*) as total
        FROM pecas
        GROUP BY label
        ORDER BY total DESC
    ''').fetchall()

    # 3. Dados para o gráfico de barras
    bar_chart_stats = db.execute('''
        SELECT strftime('%Y-%m', datetime(p.data_criacao, 'localtime')) as mes, i.nome as peca_nome, COUNT(i.id) as total
        FROM pedidos p JOIN pecas i ON p.id = i.pedido_id
        GROUP BY mes, peca_nome
        ORDER BY mes, peca_nome
    ''').fetchall()
    db.close()

    # --- Criação dos Gráficos como Imagens ---
    pie_chart_img = create_pie_chart_image(pie_chart_data)
    bar_chart_img = create_bar_chart_image(bar_chart_stats)

    # --- Criação do PDF ---
    pdf = FPDF()
    pdf.add_page()

    # Título
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Relatório de Peças e Estatísticas', 0, 1, 'C')
    pdf.ln(10)

    # Gráficos
    if pie_chart_img:
        pdf.image(pie_chart_img, x=10, y=30, w=90)
    if bar_chart_img:
        pdf.image(bar_chart_img, x=110, y=30, w=90)
    pdf.ln(80) # Espaço para os gráficos

    # Tabela de Peças
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, 'Lista de Peças', 0, 1, 'L')
    pdf.ln(5)

    pdf.set_font('Arial', 'B', 10)
    pdf.cell(30, 10, 'Pedido', 1)
    pdf.cell(60, 10, 'Defeito', 1)
    pdf.cell(40, 10, 'Peça', 1)
    pdf.cell(30, 10, 'Técnico', 1)
    pdf.cell(30, 10, 'Data', 1)
    pdf.ln()

    pdf.set_font('Arial', '', 10)
    for peca in pecas_list:
        pdf.cell(30, 10, str(peca['numero_pedido']), 1)
        pdf.cell(60, 10, peca['defeito'], 1)
        pdf.cell(40, 10, peca['nome'], 1)
        pdf.cell(30, 10, peca['tecnico_nome'], 1)
        pdf.cell(30, 10, peca['data_formatada'], 1)
        pdf.ln()

    # --- Finalização ---
    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)

    return Response(
        output,
        mimetype='application/pdf',
        headers={"Content-Disposition": "attachment;filename=relatorio.pdf"}
    )

def create_pie_chart_image(data):
    if not data:
        return None

    labels = [row['label'] for row in data]
    sizes = [row['total'] for row in data]

    fig, ax = plt.subplots()
    ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
    ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
    plt.title('Peças por Defeito')

    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png')
    img_buffer.seek(0)
    plt.close(fig)
    return img_buffer

def create_bar_chart_image(data):
    if not data:
        return None

    # Pivotar os dados
    pivot_data = collections.defaultdict(lambda: collections.defaultdict(int))
    all_pecas = sorted(list(set(row['peca_nome'] for row in data)))
    all_months = sorted(list(set(row['mes'] for row in data)))
    for row in data:
        pivot_data[row['mes']][row['peca_nome']] = row['total']

    fig, ax = plt.subplots(figsize=(10, 6))

    bar_width = 0.2
    index = range(len(all_months))

    for i, peca in enumerate(all_pecas):
        peca_data = [pivot_data[month].get(peca, 0) for month in all_months]
        ax.bar([x + i * bar_width for x in index], peca_data, bar_width, label=peca)

    ax.set_xlabel('Mês')
    ax.set_ylabel('Quantidade')
    ax.set_title('Volume Mensal de Peças')
    ax.set_xticks([x + bar_width * (len(all_pecas) - 1) / 2 for x in index])
    ax.set_xticklabels(all_months)
    ax.legend()

    plt.tight_layout()

    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png')
    img_buffer.seek(0)
    plt.close(fig)
    return img_buffer

@app.route('/api/stats/pecas_por_defeito')
def pecas_por_defeito_stats():
    db = get_db()
    stats = db.execute('''
        SELECT nome, defeito, COUNT(*) as total
        FROM pecas
        GROUP BY nome, defeito
        ORDER BY total DESC
    ''').fetchall()
    db.close()

    # Formata os dados para o gráfico
    chart_data = [
        {
            'label': f"{row['nome']} - {row['defeito']}",
            'total': row['total']
        }
        for row in stats
    ]
    return jsonify(chart_data)

@app.route('/api/stats/pecas_mensal')
def pecas_mensal_stats():
    db = get_db()
    stats = db.execute('''
        SELECT 
            strftime('%Y-%m', datetime(p.data_criacao, 'localtime')) as mes,
            i.nome as peca_nome,
            COUNT(i.id) as total
        FROM pedidos p
        JOIN pecas i ON p.id = i.pedido_id
        GROUP BY mes, peca_nome
        ORDER BY mes, peca_nome
    ''').fetchall()
    db.close()

    # Processar os dados para o formato de gráfico de barras empilhadas
    data_by_month = collections.defaultdict(lambda: collections.defaultdict(int))
    all_pecas = set()
    all_months = []

    for row in stats:
        mes = row['mes']
        peca_nome = row['peca_nome']
        total = row['total']
        
        if mes not in all_months:
            all_months.append(mes)
        all_pecas.add(peca_nome)
        data_by_month[mes][peca_nome] = total

    all_months.sort()
    sorted_pecas = sorted(list(all_pecas))

    datasets = []
    for peca in sorted_pecas:
        dataset = {
            'label': peca,
            'data': [data_by_month[month][peca] for month in all_months]
        }
        datasets.append(dataset)

    chart_data = {
        'labels': all_months,
        'datasets': datasets
    }

    return jsonify(chart_data)

if __name__ == '__main__':
    # Descomente a linha abaixo apenas na primeira execução para criar o banco de dados
    # init_db()
    # NOTA: Para usar em rede local, execute server.py ao invés deste arquivo
    app.run(host='127.0.0.1', port=5000, debug=False)
